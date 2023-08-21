'''
Author: 27
LastEditors: 27
Date: 2023-08-17 15:46:30
LastEditTime: 2023-08-22 05:40:10
FilePath: /Coding-Daily/content/PythonPractice/demo/excel_remove_null_row_and_col.py
description: type some description
'''
from typing import List
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException


def read_excel_file_2_workbook() -> (Workbook, Worksheet):
    # wb = load_workbook('/Users/f27/self_repo/Coding-Daily/content/PythonPractice/demo/remove_null_row_and_col.xlsx')
    wb = load_workbook('/Users/f27/self_repo/Coding-Daily/content/PythonPractice/demo/test11.xlsx')
    ws: Worksheet = wb.active
    return wb, ws

def test11_iter():
    wb, ws = read_excel_file_2_workbook()
    # 遍历 test11.xlsx 文件,  构造一个字典
    # 拿出头两列的数据，如果 是 字符串且去掉旁边的空格
    titles = [ws.cell(row=1, column=i).value.strip() for i in range(1, 3) if isinstance(ws.cell(row=1, column=i).value, str)]
    print(titles)
    contents: List[List] = []
    # 收集前两列的数据到 contents 当两列都是空的时候跳过 第一行也跳过

    for row in ws.iter_rows():
        if row[0].row == 1:
            continue
        if all([cell.value is None or (isinstance(cell.value, str) and cell.value.strip()) == '' for cell in row]):
            continue
        # 取每行前两列的cell
        spu_id = ws.cell(row=row[0].row, column=1).value
        spu_name = ws.cell(row=row[0].row, column=2).value
        contents.append([spu_id, spu_name])
    print(contents, len(contents))



def trim_str_space():
    # 给字符串去除两边的空格
    str = '  123  '
    # print(str.strip())


# 读取excel文件，遍历第一个 sheet，去除所有空行和空列，并生成一个新的excel文件内容
def iter_excel_and_remove_null_row_and_col():
    wb, ws = read_excel_file_2_workbook()
    # 遍历行
    for row in ws.iter_rows():
        # 如果当前行每一个cell都是None 或者 空字符串，则删除当前行
        if all([cell.value is None or (isinstance(cell.value, str) and cell.value.strip()) == '' for cell in row]):
            ws.delete_rows(row[0].row, 1)
        # if all([cell.value is None for cell in row]):
            # print(row[0].row, row)
    # 遍历列
    for col in ws.iter_cols():
        # 如果当前列每一个cell都是None，则删除当前列
        if all([cell.value is None or (isinstance(cell.value, str) and cell.value.strip()) == '' for cell in col]):
            ws.delete_cols(col[0].column, 1)
    # ws assign to  wb
    wb.active = ws
    # 保存文件
    wb.save('/Users/f27/self_repo/Coding-Daily/content/PythonPractice/demo/remove_null_row_and_col.xlsx')

def test1():
    a = "1"
    b = 1
    c = [1, 2, 3]
    print(isinstance(a, str))
    print(isinstance(b, int))
    print(isinstance(a, int))
    print(isinstance(b, str))
    print(isinstance(c, list))
import re

def extract_file_format(input_string):
    # 使用正则表达式匹配文件格式字符串
    match = re.search(r'\.\w+\s*file format', input_string)
    if match:
        file_format = match.group(0).split()[0]
        return file_format
    return "未知文件类型"




def test2():
    # 使用 load_workbook 校验文件后缀，如果不是 xlsx 结尾的文件，报错
    try:
        # wb = load_workbook('/Users/f27/self_repo/Coding-Daily/content/PythonPractice/demo/发布流程 2.png')
        wb = load_workbook('/Users/f27/self_repo/Coding-Daily/content/PythonPractice/demo/t1.txt')
    except InvalidFileException as e:
        # print(str(e))
        a = extract_file_format(str(e))
        print(a)
    pass

if __name__ == "__main__":
    # iter_excel_and_remove_null_row_and_col()
    # test1()
    # test11_iter()
    test2()
    pass
